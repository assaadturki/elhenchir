# -*- coding: utf-8 -*-
from functools import wraps
from flask import Flask, render_template, redirect, url_for, flash, request, jsonify, abort, send_from_directory
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_migrate import Migrate
from datetime import datetime, timedelta
from sqlalchemy import func
import os
from dotenv import load_dotenv
import uuid
from werkzeug.utils import secure_filename

# Charger .env avant tout
load_dotenv()

from config import get_config
from models import (db, User, Plot, Crop, Expense, Revenue, Inventory, InventoryMovement,
                    Employee, Attendance, Harvest, Notification,
                    AnimalType, Animal, AnimalSoin, AnimalProduction, Alimentation)

# ========== CREATION DE L'APPLICATION ==========
app = Flask(__name__)
app.config.from_object(get_config())

db.init_app(app)
migrate = Migrate(app, db)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Veuillez vous connecter pour acceder a cette page.'
login_manager.login_message_category = 'warning'


# ========== DECORATEURS DE ROLES ==========
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            abort(403)
        return f(*args, **kwargs)
    return decorated_function

def manager_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role not in ['admin', 'manager']:
            abort(403)
        return f(*args, **kwargs)
    return decorated_function


# ========== GESTIONNAIRES D'ERREUR ==========
@app.errorhandler(403)
def forbidden(e):
    return render_template('errors/403.html'), 403


# ========== LOAD USER ==========
@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


# ========== CREATION DES TABLES + ADMIN PAR DEFAUT ==========
def init_db():
    with app.app_context():
        db.create_all()
        # Creer le dossier exports avec chemin absolu
        exports_dir = app.config['EXPORT_FOLDER']
        uploads_dir = app.config['UPLOAD_FOLDER']
        os.makedirs(exports_dir, exist_ok=True)
        os.makedirs(uploads_dir, exist_ok=True)

        # Creer admin depuis .env (pas de mot de passe hardcode)
        admin_email = app.config['ADMIN_EMAIL']
        if not User.query.filter_by(email=admin_email).first():
            admin = User(
                nom=app.config['ADMIN_NOM'],
                email=admin_email,
                role='admin'
            )
            admin.set_password(app.config['ADMIN_PASSWORD'])
            db.session.add(admin)
            db.session.commit()
            print(f"Admin cree: {admin_email}")


init_db()

# ========== UPLOAD FACTURES ==========
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def save_facture(file):
    if not file or file.filename == '' or not allowed_file(file.filename):
        return None
    ext = file.filename.rsplit('.', 1)[1].lower()
    filename = f"{uuid.uuid4().hex}.{ext}"
    factures_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads', 'factures')
    os.makedirs(factures_dir, exist_ok=True)
    file.save(os.path.join(factures_dir, filename))
    return filename

@app.route('/uploads/factures/<path:filename>')
@login_required
def serve_facture(filename):
    factures_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads', 'factures')
    return send_from_directory(factures_dir, filename)



# ========== GESTIONNAIRES D'ERREUR ==========
@app.errorhandler(404)
def page_not_found(e):
    return render_template('errors/404.html'), 404


@app.errorhandler(500)
def internal_error(e):
    db.session.rollback()
    return render_template('errors/500.html'), 500


# ========== ROUTES PRINCIPALES ==========
@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        user = User.query.filter_by(email=email).first()

        if user and user.check_password(password) and user.is_active:
            login_user(user)
            next_page = request.args.get('next')
            flash('Connexion reussie!', 'success')
            return redirect(next_page or url_for('dashboard'))
        else:
            flash('Email ou mot de passe incorrect', 'danger')

    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Deconnexion reussie', 'info')
    return redirect(url_for('login'))


@app.route('/dashboard')
@login_required
def dashboard():
    today = datetime.now().date()
    first_day = today.replace(day=1)

    monthly_expenses = db.session.query(func.sum(Expense.montant_tnd)).filter(
        Expense.date_depense >= first_day
    ).scalar() or 0

    monthly_revenues = db.session.query(func.sum(Revenue.revenu_total_tnd)).filter(
        Revenue.date_vente >= first_day
    ).scalar() or 0

    monthly_profit = float(monthly_revenues) - float(monthly_expenses)
    active_crops = Crop.query.filter(Crop.statut.in_(['planifie', 'en_cours'])).count()
    low_stock = Inventory.query.filter(Inventory.quantite <= Inventory.seuil_alerte).all()

    return render_template('dashboard.html',
                           monthly_expenses=float(monthly_expenses),
                           monthly_revenues=float(monthly_revenues),
                           monthly_profit=monthly_profit,
                           active_crops=active_crops,
                           low_stock=low_stock)


# ========== PARCELLES ==========
@app.route('/plots')
@login_required
def plots():
    all_plots = Plot.query.all()
    return render_template('plots/index.html', plots=all_plots)


@app.route('/plots/create', methods=['GET', 'POST'])
@login_required
def create_plot():
    if request.method == 'POST':
        try:
            plot = Plot(
                nom=request.form.get('nom'),
                superficie=float(request.form.get('superficie')),
                localisation=request.form.get('localisation'),
                type_sol=request.form.get('type_sol'),
                irrigation=request.form.get('irrigation'),
                statut=request.form.get('statut', 'active')
            )
            db.session.add(plot)
            db.session.commit()
            flash('Parcelle ajoutee avec succes!', 'success')
            return redirect(url_for('plots'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur: {str(e)}', 'danger')
    return render_template('plots/create.html')


@app.route('/plots/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_plot(id):
    plot = db.get_or_404(Plot, id)
    if request.method == 'POST':
        try:
            plot.nom = request.form.get('nom')
            plot.superficie = float(request.form.get('superficie'))
            plot.localisation = request.form.get('localisation')
            plot.type_sol = request.form.get('type_sol')
            plot.irrigation = request.form.get('irrigation')
            plot.statut = request.form.get('statut')
            db.session.commit()
            flash('Parcelle modifiee avec succes!', 'success')
            return redirect(url_for('plots'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur: {str(e)}', 'danger')
    return render_template('plots/edit.html', plot=plot)


@app.route('/plots/delete/<int:id>', methods=['POST'])
@login_required
def delete_plot(id):
    plot = db.get_or_404(Plot, id)
    try:
        db.session.delete(plot)
        db.session.commit()
        flash('Parcelle supprimee avec succes!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Impossible de supprimer: {str(e)}', 'danger')
    return redirect(url_for('plots'))


# ========== CULTURES ==========
@app.route('/crops')
@login_required
def crops():
    all_crops = Crop.query.all()
    return render_template('crops/index.html', crops=all_crops)


@app.route('/crops/create', methods=['GET', 'POST'])
@login_required
def create_crop():
    plots_list = Plot.query.all()
    if request.method == 'POST':
        try:
            crop = Crop(
                parcelle_id=int(request.form.get('parcelle_id')),
                nom_culture=request.form.get('nom_culture'),
                variete=request.form.get('variete'),
                date_semis=datetime.strptime(request.form.get('date_semis'), '%Y-%m-%d').date(),
                date_recolte_prevue=datetime.strptime(request.form.get('date_recolte_prevue'), '%Y-%m-%d').date()
                    if request.form.get('date_recolte_prevue') else None,
                quantite_estimee=float(request.form.get('quantite_estimee'))
                    if request.form.get('quantite_estimee') else None,
                unite=request.form.get('unite'),
                notes=request.form.get('notes')
            )
            db.session.add(crop)
            db.session.commit()
            flash('Culture ajoutee avec succes!', 'success')
            return redirect(url_for('crops'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur: {str(e)}', 'danger')
    return render_template('crops/create.html', plots=plots_list)


@app.route('/crops/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_crop(id):
    crop = db.get_or_404(Crop, id)
    plots_list = Plot.query.all()
    if request.method == 'POST':
        try:
            crop.parcelle_id = int(request.form.get('parcelle_id'))
            crop.nom_culture = request.form.get('nom_culture')
            crop.variete = request.form.get('variete')
            crop.date_semis = datetime.strptime(request.form.get('date_semis'), '%Y-%m-%d').date()
            crop.date_recolte_prevue = datetime.strptime(request.form.get('date_recolte_prevue'), '%Y-%m-%d').date() \
                if request.form.get('date_recolte_prevue') else None
            crop.quantite_estimee = float(request.form.get('quantite_estimee')) \
                if request.form.get('quantite_estimee') else None
            crop.unite = request.form.get('unite')
            crop.statut = request.form.get('statut')
            crop.notes = request.form.get('notes')
            db.session.commit()
            flash('Culture modifiee avec succes!', 'success')
            return redirect(url_for('crops'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur: {str(e)}', 'danger')
    return render_template('crops/edit.html', crop=crop, plots=plots_list)


# ========== DEPENSES ==========
@app.route('/expenses')
@login_required
def expenses():
    all_expenses = Expense.query.order_by(Expense.date_depense.desc()).all()
    return render_template('expenses/index.html', expenses=all_expenses)


@app.route('/expenses/create', methods=['GET', 'POST'])
@login_required
def create_expense():
    plots_list = Plot.query.all()
    categories = ['Semences', 'Engrais', 'Eau', 'Carburant', 'Salaires',
                  'Transport', 'Reparation materiel', 'Electricite',
                  'Produits veterinaires', 'Nourriture animale', 'Loyer', 'Assurances', 'Autre']

    if request.method == 'POST':
        try:
            facture_fichier = save_facture(request.files.get('facture'))
            expense = Expense(
                categorie=request.form.get('categorie'),
                montant_tnd=float(request.form.get('montant_tnd')),
                date_depense=datetime.strptime(request.form.get('date_depense'), '%Y-%m-%d').date(),
                description=request.form.get('description'),
                parcelle_id=int(request.form.get('parcelle_id')) if request.form.get('parcelle_id') else None,
                fournisseur=request.form.get('fournisseur'),
                reference=request.form.get('reference'),
                facture_fichier=facture_fichier
            )
            db.session.add(expense)
            db.session.commit()
            flash('Depense ajoutee avec succes!', 'success')
            return redirect(url_for('expenses'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur: {str(e)}', 'danger')

    return render_template('expenses/create.html', plots=plots_list, categories=categories, now=datetime.now())


# ========== REVENUS ==========
@app.route('/revenues')
@login_required
def revenues():
    all_revenues = Revenue.query.order_by(Revenue.date_vente.desc()).all()
    return render_template('revenues/index.html', revenues=all_revenues)


@app.route('/revenues/create', methods=['GET', 'POST'])
@login_required
def create_revenue():
    crops_list = Crop.query.all()

    if request.method == 'POST':
        try:
            quantite = float(request.form.get('quantite_vendue'))
            prix = float(request.form.get('prix_unitaire_tnd'))
            total = quantite * prix

            revenue = Revenue(
                culture_id=int(request.form.get('culture_id')) if request.form.get('culture_id') else None,
                quantite_vendue=quantite,
                prix_unitaire_tnd=prix,
                revenu_total_tnd=total,
                date_vente=datetime.strptime(request.form.get('date_vente'), '%Y-%m-%d').date(),
                client=request.form.get('client'),
                facture_numero=request.form.get('facture_numero'),
                notes=request.form.get('notes'),
                source='culture'
            )
            db.session.add(revenue)
            db.session.commit()
            flash('Vente enregistree avec succes!', 'success')
            return redirect(url_for('revenues'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur: {str(e)}', 'danger')

    return render_template('revenues/create.html', crops=crops_list, now=datetime.now())


@app.route('/revenues/delete/<int:id>', methods=['POST'])
@login_required
def delete_revenue(id):
    revenue = db.get_or_404(Revenue, id)
    try:
        db.session.delete(revenue)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/revenues/delete/all', methods=['POST'])
@login_required
def delete_all_revenues():
    # FIX: Revenu -> Revenue (typo original)
    try:
        Revenue.query.filter(
            (Revenue.culture_id == None) |
            (Revenue.client == "Vente huile d'olive") |
            (Revenue.revenu_total_tnd == 0)
        ).delete(synchronize_session=False)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ========== STOCK ==========
@app.route('/inventory')
@login_required
def inventory():
    items = Inventory.query.all()
    return render_template('inventory/index.html', items=items)


@app.route('/inventory/create', methods=['GET', 'POST'])
@login_required
def create_inventory():
    categories = ['Engrais', 'Semences', 'Produits chimiques', 'Nourriture animale',
                  'Pieces materiel', 'Outils', 'Autre']

    if request.method == 'POST':
        try:
            item = Inventory(
                produit=request.form.get('produit'),
                categorie=request.form.get('categorie'),
                quantite=float(request.form.get('quantite')),
                unite=request.form.get('unite'),
                seuil_alerte=float(request.form.get('seuil_alerte')) if request.form.get('seuil_alerte') else 10,
                prix_unitaire=float(request.form.get('prix_unitaire')) if request.form.get('prix_unitaire') else None,
                emplacement=request.form.get('emplacement'),
                fournisseur=request.form.get('fournisseur')
            )
            db.session.add(item)
            db.session.flush()

            movement = InventoryMovement(
                produit_id=item.id,
                type_mouvement='entree',
                quantite=float(request.form.get('quantite')),
                raison='Creation initiale du stock',
                utilisateur_id=current_user.id
            )
            db.session.add(movement)
            db.session.commit()
            flash('Produit ajoute au stock avec succes!', 'success')
            return redirect(url_for('inventory'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur: {str(e)}', 'danger')

    return render_template('inventory/create.html', categories=categories)


@app.route('/inventory/movement/<int:id>', methods=['POST'])
@login_required
def add_movement(id):
    item = db.get_or_404(Inventory, id)
    try:
        type_mvt = request.form.get('type_mouvement')
        quantite = float(request.form.get('quantite'))
        raison = request.form.get('raison')

        if type_mvt == 'sortie' and quantite > item.quantite:
            flash('Quantite insuffisante en stock!', 'danger')
            return redirect(url_for('inventory'))

        if type_mvt == 'entree':
            item.quantite += quantite
        else:
            item.quantite -= quantite

        movement = InventoryMovement(
            produit_id=item.id,
            type_mouvement=type_mvt,
            quantite=quantite,
            raison=raison,
            utilisateur_id=current_user.id
        )
        db.session.add(movement)
        db.session.commit()
        flash('Mouvement de stock enregistre avec succes!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur: {str(e)}', 'danger')
    return redirect(url_for('inventory'))


# ========== EMPLOYES ==========
@app.route('/employees')
@login_required
def employees():
    all_employees = Employee.query.all()
    return render_template('employees/index.html', employees=all_employees)


@app.route('/employees/create', methods=['GET', 'POST'])
@login_required
def create_employee():
    if request.method == 'POST':
        try:
            employee = Employee(
                nom=request.form.get('nom'),
                prenom=request.form.get('prenom'),
                telephone=request.form.get('telephone'),
                email=request.form.get('email'),
                poste=request.form.get('poste'),
                salaire_tnd=float(request.form.get('salaire_tnd')) if request.form.get('salaire_tnd') else None,
                date_embauche=datetime.strptime(request.form.get('date_embauche'), '%Y-%m-%d').date()
                    if request.form.get('date_embauche') else None,
                date_naissance=datetime.strptime(request.form.get('date_naissance'), '%Y-%m-%d').date()
                    if request.form.get('date_naissance') else None,
                adresse=request.form.get('adresse'),
                cin=request.form.get('cin'),
                notes=request.form.get('notes')
            )
            db.session.add(employee)
            db.session.commit()
            flash('Employe ajoute avec succes!', 'success')
            return redirect(url_for('employees'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur: {str(e)}', 'danger')
    return render_template('employees/create.html')


@app.route('/employees/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_employee(id):
    employee = db.get_or_404(Employee, id)
    if request.method == 'POST':
        try:
            employee.nom = request.form.get('nom')
            employee.prenom = request.form.get('prenom')
            employee.telephone = request.form.get('telephone')
            employee.email = request.form.get('email')
            employee.poste = request.form.get('poste')
            employee.salaire_tnd = float(request.form.get('salaire_tnd')) if request.form.get('salaire_tnd') else None
            employee.date_embauche = datetime.strptime(request.form.get('date_embauche'), '%Y-%m-%d').date() \
                if request.form.get('date_embauche') else None
            employee.date_naissance = datetime.strptime(request.form.get('date_naissance'), '%Y-%m-%d').date() \
                if request.form.get('date_naissance') else None
            employee.adresse = request.form.get('adresse')
            employee.cin = request.form.get('cin')
            employee.statut = request.form.get('statut')
            employee.notes = request.form.get('notes')
            db.session.commit()
            flash('Employe modifie avec succes!', 'success')
            return redirect(url_for('employees'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur: {str(e)}', 'danger')
    return render_template('employees/edit.html', employee=employee)


@app.route('/employees/attendance/<int:id>', methods=['GET', 'POST'])
@login_required
def manage_attendance(id):
    employee = db.get_or_404(Employee, id)
    if request.method == 'POST':
        try:
            date = datetime.strptime(request.form.get('date'), '%Y-%m-%d').date()
            arrivee = datetime.strptime(request.form.get('heure_arrivee'), '%H:%M').time() \
                if request.form.get('heure_arrivee') else None
            depart = datetime.strptime(request.form.get('heure_depart'), '%H:%M').time() \
                if request.form.get('heure_depart') else None

            heures = 0
            if arrivee and depart:
                delta = datetime.combine(date, depart) - datetime.combine(date, arrivee)
                heures = max(0, delta.total_seconds() / 3600)

            attendance = Attendance(
                employe_id=employee.id,
                date=date,
                heure_arrivee=arrivee,
                heure_depart=depart,
                heures_travaillees=heures,
                statut=request.form.get('statut'),
                notes=request.form.get('notes')
            )
            db.session.add(attendance)
            db.session.commit()
            flash('Presence enregistree avec succes!', 'success')
            return redirect(url_for('employees'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur: {str(e)}', 'danger')
    return render_template('employees/attendance.html', employee=employee)


# ========== RECOLTES ==========
@app.route('/harvests')
@login_required
def harvests():
    all_harvests = Harvest.query.order_by(Harvest.date_recolte.desc()).all()
    return render_template('harvests/index.html', harvests=all_harvests)


@app.route('/harvests/<int:id>')
@login_required
def harvests_detail(id):
    harvest = db.get_or_404(Harvest, id)
    return render_template('harvests/detail.html', harvest=harvest)


@app.route('/harvests/create/<int:crop_id>', methods=['GET', 'POST'])
@login_required
def harvests_create(crop_id):
    crop = db.get_or_404(Crop, crop_id)

    if request.method == 'POST':
        try:
            date_recolte = datetime.strptime(request.form.get('date_recolte'), '%Y-%m-%d').date()
            quantite = float(request.form.get('quantite_recoltee'))
            pertes = float(request.form.get('pertes')) if request.form.get('pertes') else 0

            harvest = Harvest(
                culture_id=crop.id,
                date_recolte=date_recolte,
                quantite_recoltee=quantite,
                unite=request.form.get('unite'),
                qualite=request.form.get('qualite'),
                pertes=pertes,
                main_doeuvre=int(request.form.get('main_doeuvre')) if request.form.get('main_doeuvre') else None,
                notes=request.form.get('notes')
            )
            db.session.add(harvest)

            crop.quantite_recoltee = quantite
            crop.date_recolte_reelle = date_recolte
            crop.statut = 'recolte'

            db.session.commit()
            flash('Recolte enregistree avec succes!', 'success')
            return redirect(url_for('crops'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur: {str(e)}', 'danger')

    qualites = ['Extra', 'Premiere', 'Deuxieme', 'Standard', 'Conservation']
    return render_template('harvests/create.html', crop=crop, now=datetime.now(), qualites=qualites)


@app.route('/harvests/delete/<int:id>', methods=['POST'])
@login_required
def harvests_delete(id):
    harvest = db.get_or_404(Harvest, id)
    try:
        db.session.delete(harvest)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/harvests/export')
@login_required
def harvests_export():
    from openpyxl import Workbook

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Recoltes"

        headers = ['Date recolte', 'Culture', 'Parcelle', 'Quantite', 'Unite', 'Qualite', 'Pertes (%)']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        all_harvests = Harvest.query.all()
        for row, harvest in enumerate(all_harvests, 2):
            ws.cell(row=row, column=1, value=harvest.date_recolte.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=2, value=harvest.crop.nom_culture if harvest.crop else '-')
            ws.cell(row=row, column=3, value=harvest.crop.plot.nom if harvest.crop and harvest.crop.plot else '-')
            ws.cell(row=row, column=4, value=float(harvest.quantite_recoltee))
            ws.cell(row=row, column=5, value=harvest.unite)
            ws.cell(row=row, column=6, value=harvest.qualite or '-')
            ws.cell(row=row, column=7, value=float(harvest.pertes))

        # FIX: chemin absolu pour l'export
        filename = f"recoltes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(app.config['EXPORT_FOLDER'], filename)
        wb.save(filepath)

        return jsonify({'success': True, 'filename': filename})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/exports/<path:filename>')
@login_required
def download_export(filename):
    return send_from_directory(app.config['EXPORT_FOLDER'], filename, as_attachment=True)


# ========== ELEVAGE ==========
@app.route('/livestock/dashboard')
@login_required
def livestock_dashboard():
    total_animaux = Animal.query.count()
    animaux_actifs = Animal.query.filter_by(statut='actif').count()
    animaux_vendus = Animal.query.filter_by(statut='vendu').count()
    production_mensuelle = db.session.query(func.sum(AnimalProduction.revenu_total)).filter(
        AnimalProduction.date_production >= datetime.now().replace(day=1)
    ).scalar() or 0
    soins_recents = AnimalSoin.query.order_by(AnimalSoin.date_soin.desc()).limit(10).all()

    return render_template('livestock/dashboard.html',
                           total_animaux=total_animaux,
                           animaux_actifs=animaux_actifs,
                           animaux_vendus=animaux_vendus,
                           production_mensuelle=float(production_mensuelle),
                           soins_recents=soins_recents)


@app.route('/livestock/types')
@login_required
def animal_types():
    types = AnimalType.query.all()
    return render_template('livestock/types.html', types=types)


@app.route('/livestock/types/create', methods=['GET', 'POST'])
@login_required
def create_animal_type():
    if request.method == 'POST':
        try:
            animal_type = AnimalType(
                nom=request.form.get('nom'),
                categorie=request.form.get('categorie'),
                description=request.form.get('description')
            )
            db.session.add(animal_type)
            db.session.commit()
            flash("Type d'animal ajoute avec succes!", 'success')
            return redirect(url_for('animal_types'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur: {str(e)}', 'danger')
    return render_template('livestock/create_type.html')


@app.route('/livestock/animals')
@login_required
def animals():
    all_animals = Animal.query.all()
    return render_template('livestock/animals.html', animals=all_animals)


@app.route('/livestock/animals/create', methods=['GET', 'POST'])
@login_required
def create_animal():
    types = AnimalType.query.all()
    plots_list = Plot.query.all()

    if request.method == 'POST':
        try:
            animal = Animal(
                type_id=int(request.form.get('type_id')),
                numero_identification=request.form.get('numero_identification'),
                nom=request.form.get('nom'),
                sexe=request.form.get('sexe'),
                date_naissance=datetime.strptime(request.form.get('date_naissance'), '%Y-%m-%d').date()
                    if request.form.get('date_naissance') else None,
                race=request.form.get('race'),
                couleur=request.form.get('couleur'),
                poids_naissance=float(request.form.get('poids_naissance')) if request.form.get('poids_naissance') else None,
                poids_actuel=float(request.form.get('poids_actuel')) if request.form.get('poids_actuel') else None,
                prix_achat=float(request.form.get('prix_achat')) if request.form.get('prix_achat') else None,
                date_achat=datetime.strptime(request.form.get('date_achat'), '%Y-%m-%d').date()
                    if request.form.get('date_achat') else None,
                provenance=request.form.get('provenance'),
                parcelle_id=int(request.form.get('parcelle_id')) if request.form.get('parcelle_id') else None,
                notes=request.form.get('notes')
            )
            db.session.add(animal)
            db.session.commit()
            flash('Animal ajoute avec succes!', 'success')
            return redirect(url_for('animals'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur: {str(e)}', 'danger')

    return render_template('livestock/create_animal.html', types=types, plots=plots_list)


@app.route('/livestock/animals/<int:id>')
@login_required
def animal_detail(id):
    animal = db.get_or_404(Animal, id)
    soins = AnimalSoin.query.filter_by(animal_id=id).order_by(AnimalSoin.date_soin.desc()).all()
    productions = AnimalProduction.query.filter_by(animal_id=id).order_by(AnimalProduction.date_production.desc()).all()
    return render_template('livestock/animal_detail.html', animal=animal, soins=soins, productions=productions, now=datetime.now())


@app.route('/livestock/animals/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_animal(id):
    animal = db.get_or_404(Animal, id)
    types = AnimalType.query.all()
    plots_list = Plot.query.all()

    if request.method == 'POST':
        try:
            animal.type_id = int(request.form.get('type_id'))
            animal.numero_identification = request.form.get('numero_identification')
            animal.nom = request.form.get('nom')
            animal.sexe = request.form.get('sexe')
            animal.date_naissance = datetime.strptime(request.form.get('date_naissance'), '%Y-%m-%d').date() \
                if request.form.get('date_naissance') else None
            animal.race = request.form.get('race')
            animal.couleur = request.form.get('couleur')
            animal.poids_naissance = float(request.form.get('poids_naissance')) if request.form.get('poids_naissance') else None
            animal.poids_actuel = float(request.form.get('poids_actuel')) if request.form.get('poids_actuel') else None
            animal.prix_achat = float(request.form.get('prix_achat')) if request.form.get('prix_achat') else None
            animal.date_achat = datetime.strptime(request.form.get('date_achat'), '%Y-%m-%d').date() \
                if request.form.get('date_achat') else None
            animal.provenance = request.form.get('provenance')
            animal.parcelle_id = int(request.form.get('parcelle_id')) if request.form.get('parcelle_id') else None
            animal.statut = request.form.get('statut')
            animal.notes = request.form.get('notes')
            db.session.commit()
            flash('Animal modifie avec succes!', 'success')
            return redirect(url_for('animals'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur: {str(e)}', 'danger')

    return render_template('livestock/edit_animal.html', animal=animal, types=types, plots=plots_list, now=datetime.now())


@app.route('/livestock/soins/create/<int:animal_id>', methods=['GET', 'POST'])
@login_required
def create_soin(animal_id):
    animal = db.get_or_404(Animal, animal_id)

    if request.method == 'POST':
        try:
            soin = AnimalSoin(
                animal_id=animal_id,
                date_soin=datetime.strptime(request.form.get('date_soin'), '%Y-%m-%d').date(),
                type_soin=request.form.get('type_soin'),
                description=request.form.get('description'),
                veterinaire=request.form.get('veterinaire'),
                cout_tnd=float(request.form.get('cout_tnd')) if request.form.get('cout_tnd') else None,
                produits_utilises=request.form.get('produits_utilises'),
                prochain_rappel=datetime.strptime(request.form.get('prochain_rappel'), '%Y-%m-%d').date()
                    if request.form.get('prochain_rappel') else None,
                notes=request.form.get('notes')
            )
            db.session.add(soin)
            db.session.commit()
            flash('Soin enregistre avec succes!', 'success')
            return redirect(url_for('animal_detail', id=animal_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur: {str(e)}', 'danger')

    types_soin = ['Vaccination', 'Traitement', 'Consultation', 'Operation', 'Controle sanitaire']
    return render_template('livestock/create_soin.html', animal=animal, types_soin=types_soin, now=datetime.now())


@app.route('/livestock/productions/create/<int:animal_id>', methods=['GET', 'POST'])
@login_required
def create_production(animal_id):
    animal = db.get_or_404(Animal, animal_id)

    if request.method == 'POST':
        try:
            quantite = float(request.form.get('quantite'))
            prix = float(request.form.get('prix_vente_unitaire')) if request.form.get('prix_vente_unitaire') else 0
            revenu = quantite * prix

            production = AnimalProduction(
                animal_id=animal_id,
                date_production=datetime.strptime(request.form.get('date_production'), '%Y-%m-%d').date(),
                type_produit=request.form.get('type_produit'),
                quantite=quantite,
                unite=request.form.get('unite'),
                prix_vente_unitaire=prix,
                revenu_total=revenu,
                notes=request.form.get('notes')
            )
            db.session.add(production)

            if revenu > 0:
                revenue = Revenue(
                    culture_id=None,
                    quantite_vendue=quantite,
                    prix_unitaire_tnd=prix,
                    revenu_total_tnd=revenu,
                    date_vente=datetime.now().date(),
                    client='Vente directe',
                    source='elevage',
                    notes=f'Production de {animal.nom or animal.numero_identification}: {request.form.get("type_produit")}'
                )
                db.session.add(revenue)

            db.session.commit()
            flash(f'Production enregistree! Revenu: {revenu:.3f} TND', 'success')
            return redirect(url_for('animal_detail', id=animal_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur: {str(e)}', 'danger')

    types_produit = ['Lait', 'Oeufs', 'Laine', 'Miel', 'Cuir', 'Viande']
    unites = ['L', 'kg', 'pieces', 'g']
    return render_template('livestock/create_production.html', animal=animal,
                           types_produit=types_produit, unites=unites, now=datetime.now())


@app.route('/livestock/alimentations')
@login_required
def alimentations():
    all_alim = Alimentation.query.order_by(Alimentation.date_alimentation.desc()).all()
    return render_template('livestock/alimentations.html', alimentations=all_alim)


@app.route('/livestock/alimentations/create', methods=['GET', 'POST'])
@login_required
def create_alimentation():
    animaux = Animal.query.filter_by(statut='actif').all()

    if request.method == 'POST':
        try:
            animal_id = int(request.form.get('animal_id'))
            animal = db.session.get(Animal, animal_id)

            if not animal:
                flash('Animal non trouve!', 'danger')
                return redirect(url_for('alimentations'))

            date_alimentation = datetime.strptime(request.form.get('date_alimentation'), '%Y-%m-%d').date() \
                if request.form.get('date_alimentation') else datetime.now().date()
            cout_tnd = float(request.form.get('cout_tnd')) if request.form.get('cout_tnd') else 0

            alimentation = Alimentation(
                animal_id=animal_id,
                date_alimentation=date_alimentation,
                type_nourriture=request.form.get('type_nourriture'),
                quantite=float(request.form.get('quantite')),
                unite=request.form.get('unite'),
                cout_tnd=cout_tnd,
                notes=request.form.get('notes')
            )
            db.session.add(alimentation)

            if cout_tnd > 0:
                expense = Expense(
                    categorie='Nourriture animale',
                    montant_tnd=cout_tnd,
                    date_depense=date_alimentation,
                    description=f'Alimentation: {request.form.get("type_nourriture")} - {animal.nom or animal.numero_identification or "Animal"}',
                    fournisseur='Achats alimentation'
                )
                db.session.add(expense)

            db.session.commit()
            flash(f'Alimentation enregistree avec succes! Cout: {cout_tnd:.3f} TND', 'success')
            return redirect(url_for('alimentations'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur: {str(e)}', 'danger')
            return redirect(url_for('create_alimentation'))

    return render_template('livestock/create_alimentation.html', animaux=animaux, now=datetime.now())


# ========== OLIVIERS ==========
@app.route('/oliviers')
@login_required
def oliviers():
    all_oliviers = Olivier.query.all()
    return render_template('oliviers/index.html', oliviers=all_oliviers)


@app.route('/oliviers/create', methods=['GET', 'POST'])
@login_required
def create_olivier():
    parcelles = Plot.query.all()
    varietes = ['Chemlali', 'Chetoui', 'Zarrazi', 'Oueslati', 'Sayali', 'Autre']

    if request.method == 'POST':
        try:
            olivier = Olivier(
                parcelle_id=int(request.form.get('parcelle_id')),
                numero_arbre=request.form.get('numero_arbre'),
                variete=request.form.get('variete'),
                age=int(request.form.get('age')) if request.form.get('age') else None,
                date_plantation=datetime.strptime(request.form.get('date_plantation'), '%Y-%m-%d').date()
                    if request.form.get('date_plantation') else None,
                etat_sante=request.form.get('etat_sante'),
                notes=request.form.get('notes')
            )
            db.session.add(olivier)
            db.session.commit()
            flash('Olivier ajoute avec succes!', 'success')
            return redirect(url_for('oliviers'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur: {str(e)}', 'danger')

    return render_template('oliviers/create.html', parcelles=parcelles, varietes=varietes)


@app.route('/oliviers/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_olivier(id):
    olivier = db.get_or_404(Olivier, id)
    parcelles = Plot.query.all()
    varietes = ['Chemlali', 'Chetoui', 'Zarrazi', 'Oueslati', 'Sayali', 'Autre']
    etats_sante = ['Excellent', 'Bon', 'Moyen', 'Mauvais']

    if request.method == 'POST':
        try:
            olivier.parcelle_id = int(request.form.get('parcelle_id'))
            olivier.numero_arbre = request.form.get('numero_arbre')
            olivier.variete = request.form.get('variete')
            olivier.age = int(request.form.get('age')) if request.form.get('age') else None
            olivier.date_plantation = datetime.strptime(request.form.get('date_plantation'), '%Y-%m-%d').date() \
                if request.form.get('date_plantation') else None
            olivier.statut = request.form.get('statut')
            olivier.etat_sante = request.form.get('etat_sante')
            olivier.notes = request.form.get('notes')
            db.session.commit()
            flash('Olivier modifie avec succes!', 'success')
            return redirect(url_for('oliviers'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur: {str(e)}', 'danger')

    return render_template('oliviers/edit.html', olivier=olivier, parcelles=parcelles,
                           varietes=varietes, etats_sante=etats_sante)


@app.route('/oliviers/<int:id>')
@login_required
def olivier_detail(id):
    olivier = db.get_or_404(Olivier, id)
    productions = ProductionOlive.query.filter_by(olivier_id=id).order_by(ProductionOlive.annee.desc()).all()
    return render_template('oliviers/detail.html', olivier=olivier, productions=productions)


@app.route('/oliviers/production/add/<int:olivier_id>', methods=['GET', 'POST'])
@login_required
def add_production_olive(olivier_id):
    olivier = db.get_or_404(Olivier, olivier_id)

    if request.method == 'POST':
        try:
            quantite_olives = float(request.form.get('quantite_olives'))
            taux_hutile = float(request.form.get('taux_hutile')) if request.form.get('taux_hutile') else 0
            quantite_huile = quantite_olives * (taux_hutile / 100)

            production = ProductionOlive(
                olivier_id=olivier_id,
                annee=int(request.form.get('annee')),
                quantite_olives=quantite_olives,
                taux_hutile=taux_hutile,
                quantite_huile=quantite_huile,
                qualite_huile=request.form.get('qualite_huile'),
                acidite=float(request.form.get('acidite')) if request.form.get('acidite') else None,
                date_recolte=datetime.strptime(request.form.get('date_recolte'), '%Y-%m-%d').date()
                    if request.form.get('date_recolte') else None,
                notes=request.form.get('notes')
            )
            db.session.add(production)

            if quantite_huile > 0:
                stock = StockMoulin(
                    campagne=request.form.get('annee'),
                    variete=olivier.variete,
                    quantite_litres=quantite_huile,
                    qualite=request.form.get('qualite_huile'),
                    prix_vente_litre=float(request.form.get('prix_huile')) if request.form.get('prix_huile') else None,
                    date_stock=datetime.now().date(),
                    notes=f"Production olivier {olivier.numero_arbre or ''}"
                )
                db.session.add(stock)

                prix_huile = float(request.form.get('prix_huile')) if request.form.get('prix_huile') else 0
                revenu_total = quantite_huile * prix_huile

                if revenu_total > 0:
                    revenue = Revenue(
                        culture_id=None,
                        quantite_vendue=quantite_huile,
                        prix_unitaire_tnd=prix_huile,
                        revenu_total_tnd=revenu_total,
                        date_vente=datetime.now().date(),
                        client=request.form.get('client') or "Vente huile d'olive",
                        source='olivier',
                        notes=f"Huile d'olive - {olivier.variete} - Campagne {request.form.get('annee')} - {quantite_huile:.2f} L"
                    )
                    db.session.add(revenue)

            db.session.commit()
            flash(f"Production enregistree! {quantite_huile:.2f} L d'huile", 'success')
            return redirect(url_for('olivier_detail', id=olivier_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur: {str(e)}', 'danger')

    qualites = ['Extra vierge', 'Vierge', 'Lampante']
    return render_template('oliviers/add_production.html', olivier=olivier, qualites=qualites, now=datetime.now())


@app.route('/stock/huile')
@login_required
def stock_huile():
    stock = StockMoulin.query.order_by(StockMoulin.campagne.desc()).all()
    return render_template('oliviers/stock_huile.html', stock=stock)


@app.route('/stock/huile/add', methods=['GET', 'POST'])
@login_required
def add_stock_huile():
    if request.method == 'POST':
        try:
            stock = StockMoulin(
                campagne=request.form.get('campagne'),
                variete=request.form.get('variete'),
                quantite_litres=float(request.form.get('quantite_litres')),
                qualite=request.form.get('qualite'),
                prix_vente_litre=float(request.form.get('prix_vente_litre'))
                    if request.form.get('prix_vente_litre') else None,
                date_stock=datetime.strptime(request.form.get('date_stock'), '%Y-%m-%d').date()
                    if request.form.get('date_stock') else datetime.now().date(),
                notes=request.form.get('notes')
            )
            db.session.add(stock)
            db.session.commit()
            flash("Stock d'huile ajoute avec succes!", 'success')
            return redirect(url_for('stock_huile'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur: {str(e)}', 'danger')

    return render_template('oliviers/add_stock_huile.html', now=datetime.now())


@app.route('/dashboard/oliveraie')
@login_required
def dashboard_oliveraie():
    total_oliviers = Olivier.query.count()
    total_actifs = Olivier.query.filter_by(statut='actif').count()
    production_totale = db.session.query(func.sum(ProductionOlive.quantite_huile)).scalar() or 0

    production_par_variete = db.session.query(
        Olivier.variete,
        func.sum(ProductionOlive.quantite_huile).label('total_huile')
    ).join(ProductionOlive, Olivier.id == ProductionOlive.olivier_id).group_by(Olivier.variete).all()

    return render_template('oliviers/dashboard.html',
                           total_oliviers=total_oliviers,
                           total_actifs=total_actifs,
                           production_totale=float(production_totale),
                           production_par_variete=production_par_variete)


@app.route('/oliviers/delete/<int:id>', methods=['POST'])
@login_required
def delete_olivier(id):
    olivier = db.get_or_404(Olivier, id)
    try:
        db.session.delete(olivier)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500



# ========== DEPENSES - EDIT/DELETE ==========
@app.route('/expenses/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@manager_required
def edit_expense(id):
    expense = db.get_or_404(Expense, id)
    plots_list = Plot.query.all()
    categories = ['Semences', 'Engrais', 'Eau', 'Carburant', 'Salaires', 'Transport',
                  'Reparation materiel', 'Electricite', 'Produits veterinaires',
                  'Nourriture animale', 'Loyer', 'Assurances', 'Autre']
    if request.method == 'POST':
        try:
            if request.files.get('facture') and request.files['facture'].filename:
                new_f = save_facture(request.files['facture'])
                if new_f:
                    expense.facture_fichier = new_f
            expense.categorie = request.form.get('categorie')
            expense.montant_tnd = float(request.form.get('montant_tnd'))
            expense.date_depense = datetime.strptime(request.form.get('date_depense'), '%Y-%m-%d').date()
            expense.description = request.form.get('description')
            expense.parcelle_id = int(request.form.get('parcelle_id')) if request.form.get('parcelle_id') else None
            expense.fournisseur = request.form.get('fournisseur')
            expense.reference = request.form.get('reference')
            db.session.commit()
            flash('Depense modifiee!', 'success')
            return redirect(url_for('expenses'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur: {str(e)}', 'danger')
    return render_template('expenses/edit.html', expense=expense, plots=plots_list, categories=categories)

@app.route('/expenses/delete/<int:id>', methods=['POST'])
@login_required
@manager_required
def delete_expense(id):
    expense = db.get_or_404(Expense, id)
    try:
        db.session.delete(expense)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# ========== MATERIEL AGRICOLE ==========
@app.route('/materiel')
@login_required
def materiel():
    all_materiel = Materiel.query.all()
    entretiens_proches = EntretienMateriel.query.filter(
        EntretienMateriel.date_prochain <= (datetime.now().date() + timedelta(days=30))
    ).order_by(EntretienMateriel.date_prochain).all()
    return render_template('materiel/index.html', materiels=all_materiel, entretiens_proches=entretiens_proches)

@app.route('/materiel/create', methods=['GET', 'POST'])
@login_required
@manager_required
def create_materiel():
    plots_list = Plot.query.all()
    types = ['Tracteur', 'Charrue', 'Semoir', 'Moissonneuse', 'Pompe irrigation',
             'Pulverisateur', 'Remorque', 'Motoculteur', 'Generateur', 'Autre']
    etats = [('bon', 'Bon'), ('moyen', 'Moyen'), ('mauvais', 'Mauvais'), ('hors_service', 'Hors service')]
    if request.method == 'POST':
        try:
            mat = Materiel(
                nom=request.form.get('nom'), type_materiel=request.form.get('type_materiel'),
                marque=request.form.get('marque'), modele=request.form.get('modele'),
                numero_serie=request.form.get('numero_serie') or None,
                annee_achat=int(request.form.get('annee_achat')) if request.form.get('annee_achat') else None,
                prix_achat=float(request.form.get('prix_achat')) if request.form.get('prix_achat') else None,
                etat=request.form.get('etat', 'bon'),
                parcelle_id=int(request.form.get('parcelle_id')) if request.form.get('parcelle_id') else None,
                localisation=request.form.get('localisation'), notes=request.form.get('notes')
            )
            db.session.add(mat)
            db.session.flush()
            prix_achat = float(request.form.get('prix_achat')) if request.form.get('prix_achat') else 0
            if prix_achat > 0:
                db.session.add(CoutMateriel(materiel_id=mat.id, type_cout='achat',
                    montant_tnd=prix_achat, date_cout=datetime.now().date(),
                    description=f"Achat: {mat.nom}"))
            db.session.commit()
            flash('Materiel ajoute!', 'success')
            return redirect(url_for('materiel'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur: {str(e)}', 'danger')
    return render_template('materiel/create.html', plots=plots_list, types=types, etats=etats, now=datetime.now())

@app.route('/materiel/<int:id>')
@login_required
def materiel_detail(id):
    mat = db.get_or_404(Materiel, id)
    entretiens = EntretienMateriel.query.filter_by(materiel_id=id).order_by(EntretienMateriel.date_entretien.desc()).all()
    pannes = PanneMateriel.query.filter_by(materiel_id=id).order_by(PanneMateriel.date_panne.desc()).all()
    couts = CoutMateriel.query.filter_by(materiel_id=id).order_by(CoutMateriel.date_cout.desc()).all()
    return render_template('materiel/detail.html', mat=mat, entretiens=entretiens, pannes=pannes,
                           couts=couts, total_couts=sum(float(c.montant_tnd) for c in couts), now=datetime.now())

@app.route('/materiel/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@manager_required
def edit_materiel(id):
    mat = db.get_or_404(Materiel, id)
    plots_list = Plot.query.all()
    types = ['Tracteur', 'Charrue', 'Semoir', 'Moissonneuse', 'Pompe irrigation',
             'Pulverisateur', 'Remorque', 'Motoculteur', 'Generateur', 'Autre']
    etats = [('bon', 'Bon'), ('moyen', 'Moyen'), ('mauvais', 'Mauvais'), ('hors_service', 'Hors service')]
    if request.method == 'POST':
        try:
            mat.nom = request.form.get('nom')
            mat.type_materiel = request.form.get('type_materiel')
            mat.marque = request.form.get('marque')
            mat.modele = request.form.get('modele')
            mat.numero_serie = request.form.get('numero_serie') or None
            mat.annee_achat = int(request.form.get('annee_achat')) if request.form.get('annee_achat') else None
            mat.prix_achat = float(request.form.get('prix_achat')) if request.form.get('prix_achat') else None
            mat.etat = request.form.get('etat')
            mat.parcelle_id = int(request.form.get('parcelle_id')) if request.form.get('parcelle_id') else None
            mat.localisation = request.form.get('localisation')
            mat.notes = request.form.get('notes')
            db.session.commit()
            flash('Materiel modifie!', 'success')
            return redirect(url_for('materiel_detail', id=id))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur: {str(e)}', 'danger')
    return render_template('materiel/edit.html', mat=mat, plots=plots_list, types=types, etats=etats)

@app.route('/materiel/delete/<int:id>', methods=['POST'])
@login_required
@admin_required
def delete_materiel(id):
    mat = db.get_or_404(Materiel, id)
    try:
        db.session.delete(mat)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/materiel/entretien/add/<int:materiel_id>', methods=['GET', 'POST'])
@login_required
@manager_required
def add_entretien(materiel_id):
    mat = db.get_or_404(Materiel, materiel_id)
    types_entretien = ['Vidange', 'Revision generale', 'Graissage', 'Remplacement filtre',
                       'Remplacement courroie', 'Verification electrique', 'Autre']
    if request.method == 'POST':
        try:
            cout = float(request.form.get('cout_tnd')) if request.form.get('cout_tnd') else 0
            date_e = datetime.strptime(request.form.get('date_entretien'), '%Y-%m-%d').date()
            db.session.add(EntretienMateriel(
                materiel_id=materiel_id, type_entretien=request.form.get('type_entretien'),
                date_entretien=date_e,
                date_prochain=datetime.strptime(request.form.get('date_prochain'), '%Y-%m-%d').date() if request.form.get('date_prochain') else None,
                description=request.form.get('description'), technicien=request.form.get('technicien'),
                cout_tnd=cout, pieces_changees=request.form.get('pieces_changees'), notes=request.form.get('notes')
            ))
            if cout > 0:
                db.session.add(CoutMateriel(materiel_id=materiel_id, type_cout='reparation',
                    montant_tnd=cout, date_cout=date_e,
                    description=f"Entretien: {request.form.get('type_entretien')}"))
                db.session.add(Expense(categorie='Reparation materiel', montant_tnd=cout, date_depense=date_e,
                    description=f"Entretien {mat.nom}: {request.form.get('type_entretien')}",
                    fournisseur=request.form.get('technicien')))
            db.session.commit()
            flash('Entretien enregistre!', 'success')
            return redirect(url_for('materiel_detail', id=materiel_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur: {str(e)}', 'danger')
    return render_template('materiel/add_entretien.html', mat=mat, types_entretien=types_entretien, now=datetime.now())

@app.route('/materiel/panne/add/<int:materiel_id>', methods=['GET', 'POST'])
@login_required
@manager_required
def add_panne(materiel_id):
    mat = db.get_or_404(Materiel, materiel_id)
    if request.method == 'POST':
        try:
            cout = float(request.form.get('cout_reparation')) if request.form.get('cout_reparation') else 0
            date_p = datetime.strptime(request.form.get('date_panne'), '%Y-%m-%d').date()
            statut = request.form.get('statut', 'en_cours')
            panne = PanneMateriel(
                materiel_id=materiel_id, date_panne=date_p,
                date_reparation=datetime.strptime(request.form.get('date_reparation'), '%Y-%m-%d').date() if request.form.get('date_reparation') else None,
                description=request.form.get('description'), cause=request.form.get('cause'),
                technicien=request.form.get('technicien'), cout_reparation=cout,
                statut=statut, notes=request.form.get('notes')
            )
            db.session.add(panne)
            if statut == 'en_cours':
                mat.etat = 'hors_service'
            elif statut == 'repare':
                mat.etat = 'bon'
            if cout > 0:
                db.session.add(CoutMateriel(materiel_id=materiel_id, type_cout='reparation',
                    montant_tnd=cout, date_cout=date_p,
                    description=f"Reparation: {request.form.get('description')[:50]}"))
                db.session.add(Expense(categorie='Reparation materiel', montant_tnd=cout, date_depense=date_p,
                    description=f"Reparation {mat.nom}: {request.form.get('description')[:100]}",
                    fournisseur=request.form.get('technicien')))
            db.session.commit()
            flash('Panne enregistree!', 'success')
            return redirect(url_for('materiel_detail', id=materiel_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur: {str(e)}', 'danger')
    return render_template('materiel/add_panne.html', mat=mat, now=datetime.now())

@app.route('/materiel/cout/add/<int:materiel_id>', methods=['GET', 'POST'])
@login_required
@manager_required
def add_cout_materiel(materiel_id):
    mat = db.get_or_404(Materiel, materiel_id)
    types_cout = [('carburant', 'Carburant'), ('reparation', 'Reparation'),
                  ('piece', 'Piece de rechange'), ('assurance', 'Assurance'), ('autre', 'Autre')]
    if request.method == 'POST':
        try:
            montant = float(request.form.get('montant_tnd'))
            type_cout = request.form.get('type_cout')
            date_c = datetime.strptime(request.form.get('date_cout'), '%Y-%m-%d').date()
            db.session.add(CoutMateriel(materiel_id=materiel_id, type_cout=type_cout,
                montant_tnd=montant, date_cout=date_c,
                description=request.form.get('description'), notes=request.form.get('notes')))
            cat_map = {'carburant': 'Carburant', 'reparation': 'Reparation materiel',
                       'piece': 'Reparation materiel', 'assurance': 'Assurances', 'autre': 'Autre'}
            db.session.add(Expense(categorie=cat_map.get(type_cout, 'Autre'), montant_tnd=montant,
                date_depense=date_c,
                description=f"{type_cout.capitalize()} - {mat.nom}: {request.form.get('description') or ''}"))
            db.session.commit()
            flash('Cout enregistre!', 'success')
            return redirect(url_for('materiel_detail', id=materiel_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur: {str(e)}', 'danger')
    return render_template('materiel/add_cout.html', mat=mat, types_cout=types_cout, now=datetime.now())

# ========== LANCEMENT ==========
if __name__ == '__main__':
    env = os.environ.get('FLASK_ENV', 'development')
    debug = env == 'development'
    app.run(debug=debug, host='0.0.0.0', port=5000)

# ========== MODELES MATERIEL ==========
class Materiel(db.Model):
    __tablename__ = 'materiels'
    id = db.Column(db.Integer, primary_key=True)
    nom = db.Column(db.String(100), nullable=False)
    type_materiel = db.Column(db.String(100))
    marque = db.Column(db.String(100))
    modele = db.Column(db.String(100))
    numero_serie = db.Column(db.String(100), unique=True)
    annee_achat = db.Column(db.Integer)
    prix_achat = db.Column(db.Numeric(10, 2))
    etat = db.Column(db.String(50), default='bon')
    parcelle_id = db.Column(db.Integer, db.ForeignKey('plots.id'))
    localisation = db.Column(db.String(200))
    notes = db.Column(db.Text)
    date_creation = db.Column(db.DateTime, default=datetime.utcnow)
    entretiens = db.relationship('EntretienMateriel', backref='materiel', lazy=True, cascade='all, delete-orphan')
    pannes = db.relationship('PanneMateriel', backref='materiel', lazy=True, cascade='all, delete-orphan')
    couts = db.relationship('CoutMateriel', backref='materiel', lazy=True, cascade='all, delete-orphan')

class EntretienMateriel(db.Model):
    __tablename__ = 'entretiens_materiel'
    id = db.Column(db.Integer, primary_key=True)
    materiel_id = db.Column(db.Integer, db.ForeignKey('materiels.id'), nullable=False)
    type_entretien = db.Column(db.String(100))
    date_entretien = db.Column(db.Date, nullable=False)
    date_prochain = db.Column(db.Date)
    description = db.Column(db.Text)
    technicien = db.Column(db.String(200))
    cout_tnd = db.Column(db.Numeric(10, 2))
    pieces_changees = db.Column(db.Text)
    notes = db.Column(db.Text)
    date_creation = db.Column(db.DateTime, default=datetime.utcnow)

class PanneMateriel(db.Model):
    __tablename__ = 'pannes_materiel'
    id = db.Column(db.Integer, primary_key=True)
    materiel_id = db.Column(db.Integer, db.ForeignKey('materiels.id'), nullable=False)
    date_panne = db.Column(db.Date, nullable=False)
    date_reparation = db.Column(db.Date)
    description = db.Column(db.Text, nullable=False)
    cause = db.Column(db.String(200))
    technicien = db.Column(db.String(200))
    cout_reparation = db.Column(db.Numeric(10, 2))
    statut = db.Column(db.String(50), default='en_cours')
    notes = db.Column(db.Text)
    date_creation = db.Column(db.DateTime, default=datetime.utcnow)

class CoutMateriel(db.Model):
    __tablename__ = 'couts_materiel'
    id = db.Column(db.Integer, primary_key=True)
    materiel_id = db.Column(db.Integer, db.ForeignKey('materiels.id'), nullable=False)
    type_cout = db.Column(db.String(50))
    montant_tnd = db.Column(db.Numeric(10, 2), nullable=False)
    date_cout = db.Column(db.Date, nullable=False)
    description = db.Column(db.String(200))
    notes = db.Column(db.Text)
    date_creation = db.Column(db.DateTime, default=datetime.utcnow)
